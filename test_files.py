import csv
import zipfile
from io import TextIOWrapper

from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader

from constants import FILES_DIR


def test_pdf_file():
    with zipfile.ZipFile(f"{FILES_DIR}/archive.zip") as zf:
        with zf.open("file_pdf.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            number_of_pages = len(reader.pages)
            page = reader.pages[0]
            text = page.extract_text().strip()
            assert number_of_pages == 1
            assert text == "Hello, world!"

def test_xlsx_file():
    with zipfile.ZipFile(f"{FILES_DIR}/archive.zip") as zf:
        with zf.open("file_xlsx.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            len_rows = 0
            for _ in sheet.iter_rows():
                len_rows += 1
            text = sheet.cell(row=1, column=1).value
            assert len_rows == 1
            assert text == "Hello, world!"

def test_csv_file():
    with zipfile.ZipFile(f"{FILES_DIR}/archive.zip") as zf:
        with zf.open("file_csv.csv") as csv_file:
            reader = list(csv.reader(TextIOWrapper(csv_file)))
            first_row = reader[0]
            text = first_row[0]
            number_of_rows = 0
            for _ in reader:
                number_of_rows += 1
            assert number_of_rows == 1
            assert text == "Hello world!"